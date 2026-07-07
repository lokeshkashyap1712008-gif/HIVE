"""
HIVE — Report Agent Worker
Generates Excel/CSV reports, sends notifications, formats output.
"""

import os
import json
import csv
import logging
import re
from datetime import datetime

logger = logging.getLogger(__name__)

# Default save location
ONE_DRIVE_DESKTOP = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
FALLBACK_DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
DESKTOP_PATH = ONE_DRIVE_DESKTOP if os.path.exists(ONE_DRIVE_DESKTOP) else FALLBACK_DESKTOP
HIVE_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), ".hive", "output")


class ReportAgent:
    @staticmethod
    async def run(description: str, context: dict = None) -> dict:
        description_lower = description.lower()
        context = context or {}

        try:
            if any(word in description_lower for word in ["email", "mail", "smtp", "send to"]):
                return await _send_email(description, context)
            elif any(word in description_lower for word in ["slack", "discord", "webhook"]):
                return await _send_webhook(description, context)
            elif any(word in description_lower for word in ["excel", "xlsx", "spreadsheet", "table"]):
                return await _create_excel(description, context)
            elif any(word in description_lower for word in ["csv"]):
                return await _create_csv(description, context)
            elif any(word in description_lower for word in ["pdf", "report", "generate", "document"]):
                return await _generate_report(description, context)
            elif any(word in description_lower for word in ["summarize", "summary", "overview"]):
                return await _summarize(description, context)
            else:
                return await _format_output(description, context)

        except Exception as e:
            logger.error(f"[ReportAgent] Error: {e}")
            return {"status": "error", "error": str(e)}


async def _create_excel(description: str, context: dict) -> dict:
    """Create a formatted Excel file from data in context or description."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"status": "error", "error": "openpyxl not installed. Run: pip install openpyxl"}

    # Extract data from context or parse from description
    data = context.get("data", [])
    if not data:
        # Try to extract data from description
        data = _extract_data_from_text(description)

    if not data:
        return {"status": "error", "error": "No data found to create Excel file"}

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Business Data"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Get headers from first row
    if data and isinstance(data[0], dict):
        headers = list(data[0].keys())
    else:
        headers = ["Name", "Type", "Location", "Phone", "Website", "Reviews"]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "") if isinstance(row_data, dict) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Alternate row coloring
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(data) + 2):
            try:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Add title row at top
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value="HIVE Business Report")
    title_cell.font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Add timestamp
    ws.insert_rows(2)
    timestamp_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    timestamp_cell.font = Font(italic=True, color="808080")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # Save file
    os.makedirs(HIVE_OUTPUT_DIR, exist_ok=True)
    filename = f"business_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(HIVE_OUTPUT_DIR, filename)

    # Also save to Desktop for easy access
    desktop_filepath = os.path.join(DESKTOP_PATH, filename)

    try:
        wb.save(filepath)
        wb.save(desktop_filepath)
        return {
            "status": "success",
            "file": filepath,
            "desktop_file": desktop_filepath,
            "rows": len(data),
            "columns": len(headers),
            "message": f"Excel file saved to: {desktop_filepath}",
        }
    except Exception as e:
        return {"status": "error", "error": f"Failed to save Excel file: {str(e)}"}


async def _create_csv(description: str, context: dict) -> dict:
    """Create a CSV file from data."""
    data = context.get("data", [])
    if not data:
        data = _extract_data_from_text(description)

    if not data:
        return {"status": "error", "error": "No data found to create CSV file"}

    os.makedirs(HIVE_OUTPUT_DIR, exist_ok=True)
    filename = f"business_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    filepath = os.path.join(HIVE_OUTPUT_DIR, filename)
    desktop_filepath = os.path.join(DESKTOP_PATH, filename)

    if data and isinstance(data[0], dict):
        headers = list(data[0].keys())
    else:
        headers = ["Name", "Type", "Location", "Phone"]

    try:
        for path in [filepath, desktop_filepath]:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.DictWriter(f, fieldnames=headers)
                writer.writeheader()
                writer.writerows(data)

        return {
            "status": "success",
            "file": filepath,
            "desktop_file": desktop_filepath,
            "rows": len(data),
            "message": f"CSV file saved to: {desktop_filepath}",
        }
    except Exception as e:
        return {"status": "error", "error": f"Failed to save CSV file: {str(e)}"}


def _extract_data_from_text(text: str) -> list[dict]:
    """Try to extract structured data from text description."""
    data = []

    # Look for JSON-like data in the text
    json_match = re.search(r'\[.*?\]', text, re.DOTALL)
    if json_match:
        try:
            parsed = json.loads(json_match.group(0))
            if isinstance(parsed, list):
                return parsed
        except:
            pass

    # Look for table-like data
    lines = text.split('\n')
    headers = None
    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Check if line looks like a header row
        if '|' in line or '\t' in line:
            separator = '|' if '|' in line else '\t'
            parts = [p.strip() for p in line.split(separator) if p.strip()]
            if len(parts) >= 2 and not headers:
                headers = parts
                continue

        # Check if line looks like data
        if headers and ('|' in line or '\t' in line):
            separator = '|' if '|' in line else '\t'
            parts = [p.strip() for p in line.split(separator) if p.strip()]
            if len(parts) == len(headers):
                row = dict(zip(headers, parts))
                data.append(row)

    return data


async def _send_email(description: str, context: dict) -> dict:
    email_match = re.search(r"[\w.+-]+@[\w-]+\.[\w.-]+", description)
    if not email_match:
        return {"status": "skipped", "reason": "No email address found in task description"}

    to_email = email_match.group(0)
    subject_match = re.search(r'subject[:\s]+["\'"]?([^\n"\']+)["\'"]?', description, re.IGNORECASE)
    subject = subject_match.group(1).strip() if subject_match else "HIVE Report"

    return {
        "status": "would_send",
        "to": to_email,
        "subject": subject,
        "tip": "Configure SMTP server in settings for actual sending",
    }


async def _send_webhook(description: str, context: dict) -> dict:
    webhook_match = re.search(r"https?://hooks\.(slack|discord)\.com/[^\s]+", description)
    if not webhook_match:
        webhook_match = re.search(r"https?://[^\s]+webhook[^\s]+", description, re.IGNORECASE)

    if not webhook_match:
        return {
            "status": "would_send",
            "webhook_url": "not_detected",
            "tip": "Provide a Slack/Discord webhook URL in the task description",
        }

    return {
        "status": "would_send",
        "webhook_url": webhook_match.group(0)[:50] + "...",
        "tip": "Configure webhook in settings for actual delivery",
    }


async def _generate_report(description: str, context: dict) -> dict:
    title_match = re.search(r'title[:\s]+["\'"]?([^\n"\']+)["\'"]?', description, re.IGNORECASE)
    title = title_match.group(1).strip() if title_match else "HIVE Report"

    return {
        "status": "ready_to_generate",
        "title": title,
        "format": "pdf",
        "tip": "PDF generation requires reportlab/weasyprint to be installed",
    }


async def _summarize(description: str, context: dict) -> dict:
    from hive.core.llm_router import chat

    messages = [
        {"role": "system", "content": "Summarize the following content in 3-5 bullet points."},
        {"role": "user", "content": description},
    ]
    result = await chat(messages, quality=False)
    return {
        "status": "success",
        "summary": result["content"],
        "source_length": len(description),
    }


async def _format_output(description: str, context: dict) -> dict:
    data = {"raw_input_length": len(description), "input_preview": description[:200]}
    return {"status": "success", "formatted": data}


async def run(description: str, context: dict = None) -> dict:
    return await ReportAgent.run(description, context)
