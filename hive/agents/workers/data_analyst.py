"""
HIVE — Data Analyst Agent
CSV/JSON/Excel analysis with statistical computation and chart generation.
Now also creates Excel files from search results.
"""

import os
import json
import csv
import logging
from typing import Optional

logger = logging.getLogger(__name__)

ONE_DRIVE_DESKTOP = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
FALLBACK_DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
DESKTOP_PATH = ONE_DRIVE_DESKTOP if os.path.exists(ONE_DRIVE_DESKTOP) else FALLBACK_DESKTOP
HIVE_OUTPUT_DIR = os.path.join(os.path.expanduser("~"), ".hive", "output")


def _load_data(file_path: str) -> tuple[list[dict], list[str]]:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            cols = reader.fieldnames or []
    elif ext == ".json":
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                rows = data
                cols = list(data[0].keys()) if data else []
            else:
                rows = [data]
                cols = list(data.keys())
    elif ext in (".xlsx", ".xls"):
        try:
            import pandas as pd
            df = pd.read_excel(file_path)
            rows = df.to_dict("records")
            cols = df.columns.tolist()
        except ImportError:
            return [], []
    else:
        return [], []

    return rows, cols


def _compute_stats(rows: list[dict], numeric_cols: list[str]) -> dict:
    import statistics

    if not rows:
        return {}

    numeric_data = {}
    str_cols = []

    for col in list(rows[0].keys()):
        vals = []
        for row in rows:
            try:
                vals.append(float(row.get(col, 0)))
                numeric_data[col] = vals
            except (ValueError, TypeError):
                str_cols.append(col)

    stats = {}
    for col, vals in numeric_data.items():
        if not vals:
            continue
        stats[col] = {
            "count": len(vals),
            "mean": round(statistics.mean(vals), 4),
            "median": round(statistics.median(vals), 4),
            "stdev": round(statistics.stdev(vals) if len(vals) > 1 else 0, 4),
            "min": round(min(vals), 4),
            "max": round(max(vals), 4),
        }

    return {"numeric": stats, "row_count": len(rows), "col_count": len(rows[0]) if rows else 0}


def _create_excel_from_data(data: list[dict], title: str = "Business Report") -> dict:
    """Create a formatted Excel file from a list of dictionaries."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"status": "error", "error": "openpyxl not installed"}

    if not data:
        return {"status": "error", "error": "No data to create Excel file"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Get headers
    if isinstance(data[0], dict):
        headers = list(data[0].keys())
    else:
        headers = ["Column " + str(i) for i in range(len(data[0]))]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(header).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "") if isinstance(row_data, dict) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(data) + 2):
            try:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Add title
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Save
    os.makedirs(HIVE_OUTPUT_DIR, exist_ok=True)
    from datetime import datetime
    filename = f"{title.lower().replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(HIVE_OUTPUT_DIR, filename)
    desktop_filepath = os.path.join(DESKTOP_PATH, filename)

    try:
        wb.save(filepath)
        wb.save(desktop_filepath)
        return {
            "status": "success",
            "file": filepath,
            "desktop_file": desktop_filepath,
            "rows": len(data),
            "columns": len(headers),
        }
    except Exception as e:
        return {"status": "error", "error": f"Failed to save: {str(e)}"}


async def run(task: str) -> dict:
    from hive.core.llm_router import chat, QWEN_TURBO

    import re

    # Check if task contains search results (from web_scout)
    if "exa_search" in task or "results" in task:
        # Try to extract data from the task description
        try:
            # Look for JSON data in the task
            json_match = re.search(r'\[.*?\]', task, re.DOTALL)
            if json_match:
                data = json.loads(json_match.group(0))
                if isinstance(data, list) and data:
                    result = _create_excel_from_data(data, "Business Report")
                    if result.get("status") == "success":
                        return {
                            "status": "success",
                            "message": f"Excel file created: {result.get('desktop_file')}",
                            "file": result.get("desktop_file"),
                            "rows": result.get("rows"),
                            "columns": result.get("columns"),
                        }
        except:
            pass

    # Original file-based analysis
    path_match = re.search(r'[A-Za-z]:[\\\/].*?\.(csv|json|xlsx|xls)', task)
    if not path_match:
        result = await chat(
            [{"role": "system", "content": "Extract any file path that looks like a CSV, JSON, or Excel file from the text."},
             {"role": "user", "content": task}],
            model=QWEN_TURBO,
            max_tokens=128,
        )
        path_match = re.search(r'[A-Za-z]:[\\\/].*?\.(csv|json|xlsx|xls)', result["content"])

    if not path_match:
        return {"status": "error", "message": "No data file found in task"}

    file_path = path_match.group(0)
    if not os.path.exists(file_path):
        return {"status": "error", "message": f"File not found: {file_path}"}

    rows, cols = _load_data(file_path)
    if not rows:
        return {"status": "error", "message": f"Could not load data from {file_path}"}

    stats = _compute_stats(rows, cols)

    insight_result = await chat(
        [{"role": "system", "content": "You are a data analyst. Analyze the data and provide key insights."},
         {"role": "user", "content": f"File: {file_path}\nRows: {len(rows)}\nColumns: {cols}\nStats: {stats}\n\nProvide 5 key insights from this data."}],
        model=QWEN_TURBO,
        max_tokens=768,
    )

    return {
        "status": "ok",
        "file": file_path,
        "rows_loaded": len(rows),
        "columns": cols,
        "statistics": stats,
        "insights": insight_result["content"],
    }
