"""Tools — file, shell, git, web, system, scheduling, crypto, archive operations."""

import os
import sys
import asyncio
import time
import json
import hashlib
import base64
import zipfile
import shutil
from pathlib import Path
from datetime import datetime, timedelta
import httpx
from rich.console import Console
from hive.permissions import check_url_access, check_path_access

console = Console()

# Jobs storage for scheduling
JOBS_FILE = os.path.join(os.path.expanduser("~"), ".hive", "jobs.json")


# Tool registry with schemas for LLM function calling
TOOLS = {
    "read_file": {
        "description": "Read the contents of a file",
        "parameters": {
            "path": {"type": "string", "description": "File path to read"},
        },
    },
    "list_directory": {
        "description": "List contents of a directory",
        "parameters": {
            "path": {"type": "string", "description": "Directory path"},
        },
    },
    "search_content": {
        "description": "Search file contents using regex pattern",
        "parameters": {
            "pattern": {"type": "string", "description": "Regex pattern"},
            "path": {"type": "string", "description": "File or directory to search"},
            "include": {"type": "string", "description": "File glob pattern (e.g. *.py)"},
        },
    },
    "edit_file": {
        "description": "Edit a file by replacing text",
        "parameters": {
            "path": {"type": "string", "description": "File path"},
            "old": {"type": "string", "description": "Text to replace"},
            "new": {"type": "string", "description": "Replacement text"},
        },
    },
    "write_file": {
        "description": "Write content to a file",
        "parameters": {
            "path": {"type": "string", "description": "File path"},
            "content": {"type": "string", "description": "Content to write"},
        },
    },
    "run_command": {
        "description": "Run a shell command",
        "parameters": {
            "command": {"type": "string", "description": "Shell command to run"},
            "workdir": {"type": "string", "description": "Working directory (optional)"},
        },
    },
    "git_status": {
        "description": "Show git status",
        "parameters": {},
    },
    "git_diff": {
        "description": "Show git diff",
        "parameters": {
            "target": {"type": "string", "description": "Diff target (optional)"},
        },
    },
    "git_commit": {
        "description": "Create a git commit",
        "parameters": {
            "message": {"type": "string", "description": "Commit message"},
            "files": {"type": "string", "description": "Files to stage (space-separated)"},
        },
    },
    "web_fetch": {
        "description": "Fetch content from a URL",
        "parameters": {
            "url": {"type": "string", "description": "URL to fetch"},
        },
    },
    "exa_search": {
        "description": "Search the web using Exa AI (Google Maps, business directories, etc.)",
        "parameters": {
            "query": {"type": "string", "description": "Search query"},
            "num_results": {"type": "integer", "description": "Number of results (default 5)"},
            "type": {"type": "string", "description": "Search type: 'neural' (semantic), 'keyword', or 'auto' (default auto)"},
            "include_domains": {"type": "string", "description": "Comma-separated domains to search (e.g. 'yelp.com,google.com')"},
            "category": {"type": "string", "description": "Category filter: 'company', 'research paper', 'news', 'linkedin profile', 'github', 'tweet', 'movie', 'song', 'personal site', 'pdf'"},
        },
    },
    "web_search": {
        "description": "Search the web for information (alias for exa_search)",
        "parameters": {
            "query": {"type": "string", "description": "Search query"},
            "num_results": {"type": "integer", "description": "Number of results (default 5)"},
        },
    },
    "create_excel": {
        "description": "Create a formatted Excel file from data",
        "parameters": {
            "data": {"type": "string", "description": "JSON array of objects to write to Excel"},
            "title": {"type": "string", "description": "Title for the Excel file"},
            "filename": {"type": "string", "description": "Filename (without extension)"},
        },
    },
    # ─── System Tools ─────────────────────────────────────────────
    "system_info": {
        "description": "Get system information (OS, Python, CPU, RAM, disk)",
        "parameters": {},
    },
    "disk_usage": {
        "description": "Check disk usage for a path",
        "parameters": {
            "path": {"type": "string", "description": "Path to check (default: current dir)"},
        },
    },
    "ping": {
        "description": "Ping a host to check connectivity",
        "parameters": {
            "host": {"type": "string", "description": "Hostname or IP to ping"},
        },
    },
    "dns_lookup": {
        "description": "Look up DNS records for a domain",
        "parameters": {
            "domain": {"type": "string", "description": "Domain to lookup"},
        },
    },
    # ─── Clipboard Tools ─────────────────────────────────────────
    "clipboard_get": {
        "description": "Get contents of clipboard",
        "parameters": {},
    },
    "clipboard_set": {
        "description": "Set clipboard contents",
        "parameters": {
            "text": {"type": "string", "description": "Text to copy to clipboard"},
        },
    },
    # ─── Environment Tools ───────────────────────────────────────
    "get_env": {
        "description": "Get an environment variable value",
        "parameters": {
            "name": {"type": "string", "description": "Environment variable name"},
        },
    },
    "list_env": {
        "description": "List all environment variables",
        "parameters": {},
    },
    # ─── Process Tools ───────────────────────────────────────────
    "list_processes": {
        "description": "List running processes",
        "parameters": {
            "filter": {"type": "string", "description": "Filter by name (optional)"},
        },
    },
    "kill_process": {
        "description": "Kill a process by PID or name",
        "parameters": {
            "target": {"type": "string", "description": "Process ID or name"},
        },
    },
    # ─── Package Tools ───────────────────────────────────────────
    "install_package": {
        "description": "Install a Python package via pip",
        "parameters": {
            "package": {"type": "string", "description": "Package name to install"},
        },
    },
    "list_packages": {
        "description": "List installed Python packages",
        "parameters": {},
    },
    # ─── Database Tools ──────────────────────────────────────────
    "sqlite_query": {
        "description": "Run a SQL query on a SQLite database",
        "parameters": {
            "db_path": {"type": "string", "description": "Path to SQLite database"},
            "query": {"type": "string", "description": "SQL query to execute"},
        },
    },
    "sqlite_tables": {
        "description": "List tables in a SQLite database",
        "parameters": {
            "db_path": {"type": "string", "description": "Path to SQLite database"},
        },
    },
    # ─── HTTP Client Tools ───────────────────────────────────────
    "http_request": {
        "description": "Make HTTP requests (GET, POST, PUT, DELETE)",
        "parameters": {
            "method": {"type": "string", "description": "HTTP method (GET, POST, PUT, DELETE)"},
            "url": {"type": "string", "description": "Request URL"},
            "headers": {"type": "string", "description": "Headers as JSON string (optional)"},
            "body": {"type": "string", "description": "Request body (optional)"},
        },
    },
    # ─── Scheduling Tools ────────────────────────────────────────
    "set_reminder": {
        "description": "Set a reminder or scheduled task",
        "parameters": {
            "message": {"type": "string", "description": "Reminder message"},
            "time": {"type": "string", "description": "When to remind (e.g., '5m', '2h', '2024-01-01 10:00')"},
        },
    },
    "list_jobs": {
        "description": "List scheduled jobs/reminders",
        "parameters": {},
    },
    "cancel_job": {
        "description": "Cancel a scheduled job by ID",
        "parameters": {
            "job_id": {"type": "string", "description": "Job ID to cancel"},
        },
    },
    # ─── Encryption Tools ────────────────────────────────────────
    "hash_text": {
        "description": "Hash text using various algorithms",
        "parameters": {
            "text": {"type": "string", "description": "Text to hash"},
            "algorithm": {"type": "string", "description": "Algorithm (md5, sha1, sha256, sha512)"},
        },
    },
    "hash_file": {
        "description": "Hash a file using various algorithms",
        "parameters": {
            "path": {"type": "string", "description": "File path to hash"},
            "algorithm": {"type": "string", "description": "Algorithm (md5, sha1, sha256, sha512)"},
        },
    },
    "base64_encode": {
        "description": "Encode text to base64",
        "parameters": {
            "text": {"type": "string", "description": "Text to encode"},
        },
    },
    "base64_decode": {
        "description": "Decode base64 to text",
        "parameters": {
            "text": {"type": "string", "description": "Base64 to decode"},
        },
    },
    # ─── Archive Tools ───────────────────────────────────────────
    "zip_files": {
        "description": "Create a zip archive from files/directories",
        "parameters": {
            "source": {"type": "string", "description": "File or directory to zip"},
            "output": {"type": "string", "description": "Output zip file path"},
        },
    },
    "unzip_file": {
        "description": "Extract a zip archive",
        "parameters": {
            "source": {"type": "string", "description": "Zip file to extract"},
            "output": {"type": "string", "description": "Output directory (optional)"},
        },
    },
    "list_zip": {
        "description": "List contents of a zip archive",
        "parameters": {
            "source": {"type": "string", "description": "Zip file to list"},
        },
    },
    # ─── Browser Tools ─────────────────────────────────────────
    "browser_open": {
        "description": "Open a URL in the browser and return page title",
        "parameters": {
            "url": {"type": "string", "description": "URL to navigate to"},
        },
    },
    "browser_click": {
        "description": "Click an element by index (from browser_inspect) or CSS selector/text",
        "parameters": {
            "index": {"type": "integer", "description": "Element index from browser_inspect (use this OR selector)"},
            "selector": {"type": "string", "description": "CSS selector or visible text (use this OR index)"},
        },
    },
    "browser_type": {
        "description": "Type into an input field by index (from browser_inspect) or CSS selector",
        "parameters": {
            "index": {"type": "integer", "description": "Element index from browser_inspect (use this OR selector)"},
            "selector": {"type": "string", "description": "CSS selector of input (use this OR index)"},
            "text": {"type": "string", "description": "Text to type (use ${key} placeholders for sensitive data)"},
            "press_enter": {"type": "boolean", "description": "Press Enter after typing (default false)"},
            "sensitive_data": {"type": "object", "description": "Sensitive data to inject. Keys are placeholders (e.g., 'password'), values are actual data."},
        },
    },
    "browser_wait": {
        "description": "Wait for an element to appear on the page",
        "parameters": {
            "selector": {"type": "string", "description": "CSS selector to wait for"},
            "timeout": {"type": "integer", "description": "Timeout in seconds (default 10)"},
        },
    },
    "browser_screenshot": {
        "description": "Take a screenshot of the current page and save to Desktop",
        "parameters": {
            "filename": {"type": "string", "description": "Filename for screenshot (default: screenshot.png)"},
        },
    },
    "browser_read": {
        "description": "Read the text content of the current page",
        "parameters": {},
    },
    "browser_inspect": {
        "description": "List all interactive elements (buttons, links, inputs) on the page with index numbers. Use index with browser_click/browser_type.",
        "parameters": {},
    },
    "browser_back": {
        "description": "Navigate back in browser history",
        "parameters": {},
    },
    "browser_close": {
        "description": "Close the browser session",
        "parameters": {},
    },
    # ─── Session Persistence Tools ─────────────────────────────
    "browser_session_save": {
        "description": "Save current browser session (cookies, localStorage) to disk for reuse",
        "parameters": {
            "name": {"type": "string", "description": "Session name (e.g., 'github', 'shopify')"},
        },
    },
    "browser_session_load": {
        "description": "Load a saved browser session (skips login on future runs)",
        "parameters": {
            "name": {"type": "string", "description": "Session name to load"},
        },
    },
    "browser_list_sessions": {
        "description": "List all saved browser sessions",
        "parameters": {},
    },
    "browser_delete_session": {
        "description": "Delete a saved browser session",
        "parameters": {
            "name": {"type": "string", "description": "Session name to delete"},
        },
    },
    # ─── Email Verification Tools ──────────────────────────────
    "browser_create_inbox": {
        "description": "Create a disposable email inbox for verification. Use the returned email in signup forms.",
        "parameters": {},
    },
    "browser_wait_for_code": {
        "description": "Wait for a verification code. Checks the disposable inbox if one was created, otherwise reads the user's personal email over IMAP (HIVE_IMAP_USER/PASSWORD or vault credential 'imap').",
        "parameters": {
            "timeout": {"type": "integer", "description": "Timeout in seconds (default 30)"},
            "sender_hint": {"type": "string", "description": "Only match emails whose sender contains this text, e.g. 'spotify' (optional)"},
        },
    },
    # ─── Browser Use Tool (Chrome Profile) ──────────────────────
    "browser_use_task": {
        "description": "Automate browser using Chrome profile with saved logins. Best for login, multi-step workflows, and sites with saved credentials.",
        "parameters": {
            "task": {"type": "string", "description": "What to do in the browser"},
            "headless": {"type": "boolean", "description": "Run invisibly (default false)"},
            "max_steps": {"type": "integer", "description": "Max steps (default 30)"},
        },
    },
    # ─── Spotify Control (uses the logged-in persistent browser) ─
    "spotify_control": {
        "description": "Control the Spotify web player (must be logged in). Actions: play_playlist (best for free accounts — plays your own playlists), list_playlists, play (search a song), pause, resume, next, previous, now_playing.",
        "parameters": {
            "action": {"type": "string", "description": "play_playlist | list_playlists | play | pause | resume | next | previous | now_playing"},
            "query": {"type": "string", "description": "Playlist or song name (play_playlist/play). Omit query to play your first sidebar playlist."},
        },
    },
    # ─── Vault Tools ────────────────────────────────────────────
    "vault_store_credential": {
        "description": "Store site login credentials in encrypted vault (never plaintext .env)",
        "parameters": {
            "site": {"type": "string", "description": "Site domain e.g. github.com"},
            "username": {"type": "string", "description": "Email or username"},
            "password": {"type": "string", "description": "Password"},
        },
    },
    "vault_list_credentials": {
        "description": "List stored credentials (no passwords shown)",
        "parameters": {},
    },
    "vault_store_card": {
        "description": "Store payment card in encrypted vault",
        "parameters": {
            "label": {"type": "string", "description": "Card label e.g. 'personal'"},
            "number": {"type": "string", "description": "Card number"},
            "expiry": {"type": "string", "description": "Expiry MM/YY"},
            "cvv": {"type": "string", "description": "CVV"},
            "name": {"type": "string", "description": "Name on card"},
            "billing_zip": {"type": "string", "description": "Billing ZIP (optional)"},
        },
    },
    "vault_list_cards": {
        "description": "List stored payment cards (last 4 digits only)",
        "parameters": {},
    },
    # ─── Signup & Checkout ──────────────────────────────────────
    "browser_signup": {
        "description": "Create a new account on a website: fill signup form, handle email verification, save credentials to vault",
        "parameters": {
            "task": {"type": "string", "description": "Signup task description"},
            "url": {"type": "string", "description": "Signup page URL"},
            "email": {"type": "string", "description": "Email to use (optional — creates disposable if omitted)"},
            "password": {"type": "string", "description": "Password (optional — auto-generated if omitted)"},
        },
    },
    "browser_checkout": {
        "description": "Run guarded browser checkout: fill cart/checkout, enter card from vault, STOP before final purchase for human confirmation",
        "parameters": {
            "task": {"type": "string", "description": "Checkout task description"},
            "url": {"type": "string", "description": "Checkout page URL (optional)"},
            "amount": {"type": "number", "description": "Expected order amount for spending cap check"},
            "card_id": {"type": "string", "description": "Vault card ID (optional — uses default card)"},
            "confirm": {"type": "boolean", "description": "Set true to place order after human review"},
        },
    },
    "browser_google_login": {
        "description": "Open visible browser for manual Google sign-in. You solve CAPTCHA/2FA; HIVE saves session as google_com.",
        "parameters": {
            "email": {"type": "string", "description": "Gmail address hint (optional)"},
        },
    },
    "browser_oauth": {
        "description": "OAuth login for github or google. Opens browser for authorization.",
        "parameters": {
            "platform": {"type": "string", "description": "github or google"},
        },
    },
}


# ─── MCP Tool Registration ──────────────────────────────────────────────

def register_mcp_tools(mcp_tools: dict):
    """Register MCP tools into the TOOLS registry.

    Args:
        mcp_tools: Dict of {hive_tool_name: hive_tool_spec} from MCP bridge
    """
    for name, spec in mcp_tools.items():
        TOOLS[name] = spec


def unregister_mcp_tools(server_name: str):
    """Remove all MCP tools from a specific server.

    Args:
        server_name: The MCP server name (e.g., 'gmail')
    """
    prefix = f"mcp_{server_name}__"
    keys_to_remove = [k for k in TOOLS if k.startswith(prefix)]
    for key in keys_to_remove:
        del TOOLS[key]


def get_mcp_tool_info(tool_name: str) -> dict | None:
    """Get MCP metadata for a tool if it's an MCP tool.

    Returns the _mcp dict with server and original tool name,
    or None if not an MCP tool.
    """
    spec = TOOLS.get(tool_name, {})
    return spec.get("_mcp")


async def execute_tool(tool_name: str, **kwargs) -> dict:
    """Execute a tool and return result."""
    start = time.time()
    try:
        detail = kwargs.get("path", kwargs.get("url", kwargs.get("command", "")))
        status_text = f"[dim]{tool_name}[/dim]"
        if detail:
            status_text += f" [dim]({detail[:40]})[/dim]"

        with console.status(status_text, spinner="dots"):
            if tool_name == "read_file":
                path = kwargs.get("path", "")
                allowed, reason = check_path_access(path, "read")
                if not allowed:
                    result = {"error": f"Path access denied: {reason}"}
                else:
                    result = await _read_file(path)
            elif tool_name == "list_directory":
                path = kwargs.get("path", "")
                allowed, reason = check_path_access(path, "read")
                if not allowed:
                    result = {"error": f"Path access denied: {reason}"}
                else:
                    result = await _list_directory(path)
            elif tool_name == "search_content":
                path = kwargs.get("path", "")
                allowed, reason = check_path_access(path, "read")
                if not allowed:
                    result = {"error": f"Path access denied: {reason}"}
                else:
                    result = await _search_content(
                        kwargs["pattern"], path, kwargs.get("include", "*")
                    )
            elif tool_name == "edit_file":
                path = kwargs.get("path", "")
                allowed, reason = check_path_access(path, "write")
                if not allowed:
                    result = {"error": f"Path access denied: {reason}"}
                else:
                    result = await _edit_file(path, kwargs["old"], kwargs["new"])
            elif tool_name == "write_file":
                path = kwargs.get("path", "")
                allowed, reason = check_path_access(path, "write")
                if not allowed:
                    result = {"error": f"Path access denied: {reason}"}
                else:
                    result = await _write_file(path, kwargs["content"])
            elif tool_name == "run_command":
                result = await _run_command(
                    kwargs["command"], kwargs.get("workdir", None)
                )
            elif tool_name == "git_status":
                result = await _run_command_list(["git", "status"])
            elif tool_name == "git_diff":
                target = kwargs.get("target", "")
                cmd = ["git", "diff"]
                if target:
                    cmd.append(target)
                result = await _run_command_list(cmd)
            elif tool_name == "git_commit":
                files = kwargs.get("files", ".")
                await _run_command_list(["git", "add", files])
                result = await _run_command_list(["git", "commit", "-m", kwargs["message"]])
            elif tool_name == "web_fetch":
                url = kwargs.get("url", "")
                allowed, reason = check_url_access(url)
                if not allowed:
                    result = {"error": f"URL access denied: {reason}"}
                else:
                    result = await _web_fetch(url)
            elif tool_name == "exa_search":
                result = await _exa_search(
                    kwargs["query"],
                    kwargs.get("num_results", 5),
                    kwargs.get("type", "auto"),
                    kwargs.get("include_domains", ""),
                    kwargs.get("category", ""),
                )
            elif tool_name == "web_search":
                # Alias for exa_search
                result = await _exa_search(
                    kwargs["query"],
                    kwargs.get("num_results", 5),
                    "auto",
                    "",
                    "",
                )
            elif tool_name == "create_excel":
                result = await _create_excel(
                    kwargs["data"],
                    kwargs.get("title", "HIVE Report"),
                    kwargs.get("filename", "report"),
                )
            # ─── System Tools ─────────────────────────────────────
            elif tool_name == "system_info":
                result = await _system_info()
            elif tool_name == "disk_usage":
                result = await _disk_usage(kwargs.get("path", "."))
            elif tool_name == "ping":
                result = await _ping(kwargs["host"])
            elif tool_name == "dns_lookup":
                result = await _dns_lookup(kwargs["domain"])
            # ─── Clipboard Tools ─────────────────────────────────
            elif tool_name == "clipboard_get":
                result = await _clipboard_get()
            elif tool_name == "clipboard_set":
                result = await _clipboard_set(kwargs["text"])
            # ─── Environment Tools ───────────────────────────────
            elif tool_name == "get_env":
                result = await _get_env(kwargs["name"])
            elif tool_name == "list_env":
                result = await _list_env()
            # ─── Process Tools ───────────────────────────────────
            elif tool_name == "list_processes":
                result = await _list_processes(kwargs.get("filter", ""))
            elif tool_name == "kill_process":
                result = await _kill_process(kwargs["target"])
            # ─── Package Tools ───────────────────────────────────
            elif tool_name == "install_package":
                result = await _install_package(kwargs["package"])
            elif tool_name == "list_packages":
                result = await _list_packages()
            # ─── Database Tools ──────────────────────────────────
            elif tool_name == "sqlite_query":
                result = await _sqlite_query(kwargs["db_path"], kwargs["query"])
            elif tool_name == "sqlite_tables":
                result = await _sqlite_tables(kwargs["db_path"])
            # ─── HTTP Client Tools ───────────────────────────────
            elif tool_name == "http_request":
                result = await _http_request(
                    kwargs["method"], kwargs["url"],
                    kwargs.get("headers", ""), kwargs.get("body", "")
                )
            # ─── Scheduling Tools ────────────────────────────────
            elif tool_name == "set_reminder":
                result = await _set_reminder(kwargs["message"], kwargs["time"])
            elif tool_name == "list_jobs":
                result = await _list_jobs()
            elif tool_name == "cancel_job":
                result = await _cancel_job(kwargs["job_id"])
            # ─── Encryption Tools ────────────────────────────────
            elif tool_name == "hash_text":
                result = await _hash_text(kwargs["text"], kwargs.get("algorithm", "sha256"))
            elif tool_name == "hash_file":
                result = await _hash_file(kwargs["path"], kwargs.get("algorithm", "sha256"))
            elif tool_name == "base64_encode":
                result = await _base64_encode(kwargs["text"])
            elif tool_name == "base64_decode":
                result = await _base64_decode(kwargs["text"])
            # ─── Archive Tools ───────────────────────────────────
            elif tool_name == "zip_files":
                result = await _zip_files(kwargs["source"], kwargs["output"])
            elif tool_name == "unzip_file":
                result = await _unzip_file(kwargs["source"], kwargs.get("output", ""))
            elif tool_name == "list_zip":
                result = await _list_zip(kwargs["source"])
            # ─── Browser Tools ─────────────────────────────────
            elif tool_name == "browser_open":
                result = await _browser_open(kwargs["url"])
            elif tool_name == "browser_click":
                result = await _browser_click(
                    selector=kwargs.get("selector", ""),
                    index=kwargs.get("index", -1),
                )
            elif tool_name == "browser_type":
                result = await _browser_type(
                    selector=kwargs.get("selector", ""),
                    text=kwargs.get("text", ""),
                    press_enter=kwargs.get("press_enter", False),
                    index=kwargs.get("index", -1),
                    sensitive_data=kwargs.get("sensitive_data"),
                )
            elif tool_name == "browser_wait":
                result = await _browser_wait(
                    kwargs["selector"],
                    kwargs.get("timeout", 10),
                )
            elif tool_name == "browser_screenshot":
                result = await _browser_screenshot(
                    kwargs.get("filename", "screenshot.png"),
                )
            elif tool_name == "browser_read":
                result = await _browser_read()
            elif tool_name == "browser_inspect":
                result = await _browser_inspect()
            elif tool_name == "browser_back":
                result = await _browser_back()
            elif tool_name == "browser_close":
                result = await _browser_close()
            elif tool_name == "browser_session_save":
                result = await _browser_session_save(kwargs["name"])
            elif tool_name == "browser_session_load":
                result = await _browser_session_load(kwargs["name"])
            elif tool_name == "browser_list_sessions":
                result = await _browser_list_sessions()
            elif tool_name == "browser_delete_session":
                result = await _browser_delete_session(kwargs["name"])
            elif tool_name == "browser_create_inbox":
                result = await _browser_create_inbox()
            elif tool_name == "browser_wait_for_code":
                result = await _browser_wait_for_code(kwargs.get("timeout", 30), kwargs.get("sender_hint", ""))

            elif tool_name == "spotify_control":
                from hive.browser import spotify
                result = await spotify.control(kwargs.get("action", ""), kwargs.get("query", ""))
            # ─── Browser Use Tool (Chrome Profile) ──────────────
            elif tool_name == "browser_use_task":
                result = await _browser_use_task(
                    kwargs["task"],
                    headless=kwargs.get("headless", False),
                    max_steps=kwargs.get("max_steps", 30),
                )
            elif tool_name == "vault_store_credential":
                result = await _vault_store_credential(
                    kwargs["site"], kwargs["username"], kwargs["password"],
                )
            elif tool_name == "vault_list_credentials":
                result = await _vault_list_credentials()
            elif tool_name == "vault_store_card":
                result = await _vault_store_card(
                    kwargs["label"], kwargs["number"], kwargs["expiry"],
                    kwargs["cvv"], kwargs.get("name", ""), kwargs.get("billing_zip", ""),
                )
            elif tool_name == "vault_list_cards":
                result = await _vault_list_cards()
            elif tool_name == "browser_signup":
                result = await _browser_signup(
                    kwargs["task"],
                    url=kwargs.get("url"),
                    email=kwargs.get("email"),
                    password=kwargs.get("password"),
                )
            elif tool_name == "browser_checkout":
                result = await _browser_checkout(
                    kwargs["task"],
                    url=kwargs.get("url"),
                    amount=kwargs.get("amount"),
                    card_id=kwargs.get("card_id"),
                    confirm=kwargs.get("confirm", False),
                )
            elif tool_name == "browser_google_login":
                result = await _browser_google_login(kwargs.get("email"))
            elif tool_name == "browser_oauth":
                result = await _browser_oauth(kwargs.get("platform", "github"))
            # ─── MCP Tools ────────────────────────────────────────
            elif tool_name.startswith("mcp_"):
                from hive.mcp.bridge import mcp_bridge
                if mcp_bridge.is_mcp_tool(tool_name):
                    result = await mcp_bridge.execute_mcp_tool(tool_name, **kwargs)
                else:
                    result = {"error": f"Unknown MCP tool: {tool_name}"}
            else:
                result = {"error": f"Unknown tool: {tool_name}"}

        duration_ms = int((time.time() - start) * 1000)
        result["duration_ms"] = duration_ms
        return result

    except Exception as e:
        duration_ms = int((time.time() - start) * 1000)
        return {"error": str(e), "duration_ms": duration_ms}


async def _read_file(path: str) -> dict:
    """Read file contents with size limit."""
    p = Path(path).resolve()
    if not p.exists():
        return {"error": f"File not found: {path}"}
    if not p.is_file():
        return {"error": f"Not a file: {path}"}
    try:
        # Limit to 100KB to prevent OOM
        max_bytes = 100 * 1024
        file_size = p.stat().st_size
        truncated = file_size > max_bytes
        
        with open(p, "r", encoding="utf-8", errors="replace") as f:
            content = f.read(max_bytes)
        
        lines = content.count("\n") + 1
        return {
            "content": content,
            "lines": lines,
            "file_size": file_size,
            "truncated": truncated,
        }
    except Exception as e:
        return {"error": str(e)}


async def _list_directory(path: str) -> dict:
    """List directory contents with limit."""
    p = Path(path).resolve()
    if not p.exists():
        return {"error": f"Directory not found: {path}"}
    if not p.is_dir():
        return {"error": f"Not a directory: {path}"}
    entries = []
    count = 0
    for entry in sorted(p.iterdir()):
        count += 1
        if count > 500:
            entries.append({"name": "... (truncated)", "type": "info"})
            break
        kind = "dir" if entry.is_dir() else "file"
        entries.append({"name": entry.name, "type": kind})
    return {"entries": entries, "count": count}


async def _search_content(pattern: str, path: str, include: str = "*") -> dict:
    """Search file contents using grep-like matching."""
    import re
    p = Path(path).resolve()
    matches = []

    if p.is_file():
        files = [p]
    elif p.is_dir():
        files = list(p.rglob(include))
    else:
        return {"error": f"Path not found: {path}"}

    regex = re.compile(pattern, re.IGNORECASE)
    for f in files[:100]:  # limit to 100 files
        try:
            text = f.read_text(encoding="utf-8", errors="replace")
            for i, line in enumerate(text.splitlines(), 1):
                if regex.search(line):
                    matches.append({
                        "file": str(f.relative_to(p) if p.is_dir() else f),
                        "line": i,
                        "content": line.strip()[:200],
                    })
        except Exception:
            continue

    return {"matches": matches[:50], "total": len(matches)}


async def _edit_file(path: str, old: str, new: str) -> dict:
    """Edit file by replacing text."""
    p = Path(path).resolve()
    if not p.exists():
        return {"error": f"File not found: {path}"}

    content = p.read_text(encoding="utf-8")
    if old not in content:
        return {"error": "Text not found in file"}

    count = content.count(old)
    new_content = content.replace(old, new, 1)
    p.write_text(new_content, encoding="utf-8")

    return {"replaced": True, "occurrences": count, "path": str(p)}


async def _write_file(path: str, content: str) -> dict:
    """Write content to file."""
    p = Path(path).resolve()
    p.parent.mkdir(parents=True, exist_ok=True)
    p.write_text(content, encoding="utf-8")
    return {"written": True, "path": str(p), "bytes": len(content.encode())}


async def _run_command(command: str, workdir: str | None = None) -> dict:
    """Run shell command asynchronously."""
    proc = await asyncio.create_subprocess_shell(
        command,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        cwd=workdir,
    )
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.communicate()
        return {"stdout": "", "stderr": "Command timed out after 30s", "exit_code": -1}
    return {
        "stdout": stdout.decode(errors="replace")[:5000],
        "stderr": stderr.decode(errors="replace")[:2000],
        "exit_code": proc.returncode,
    }


async def _run_command_list(cmd: list, workdir: str | None = None) -> dict:
    """Run command as list (no shell injection)."""
    proc = await asyncio.create_subprocess_exec(
        *cmd,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
        cwd=workdir,
    )
    try:
        stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
    except asyncio.TimeoutError:
        proc.kill()
        await proc.communicate()
        return {"stdout": "", "stderr": "Command timed out after 30s", "exit_code": -1}
    return {
        "stdout": stdout.decode(errors="replace")[:5000],
        "stderr": stderr.decode(errors="replace")[:2000],
        "exit_code": proc.returncode,
    }


async def _web_fetch(url: str) -> dict:
    """Fetch URL content."""
    async with httpx.AsyncClient(follow_redirects=True) as client:
        resp = await client.get(url, timeout=15.0)
        content_type = resp.headers.get("content-type", "")
        if "json" in content_type:
            return {"content": resp.json(), "status": resp.status_code}
        return {"content": resp.text[:10000], "status": resp.status_code}


async def _exa_search(
    query: str,
    num_results: int = 5,
    search_type: str = "auto",
    include_domains: str = "",
    category: str = "",
) -> dict:
    """Search the web using Exa AI API."""
    import os
    api_key = os.environ.get("EXA_API_KEY", "")
    if not api_key:
        return {"error": "EXA_API_KEY not set in environment. Add it to .env file."}

    url = "https://api.exa.ai/search"
    headers = {
        "x-api-key": api_key,
        "Content-Type": "application/json",
    }

    payload = {
        "query": query,
        "numResults": min(num_results, 20),
        "type": search_type,
        "contents": {
            "text": True,
        },
    }

    if include_domains:
        payload["includeDomains"] = [d.strip() for d in include_domains.split(",") if d.strip()]

    if category:
        payload["category"] = category

    async with httpx.AsyncClient() as client:
        try:
            resp = await client.post(url, json=payload, headers=headers, timeout=30.0)
            if resp.status_code != 200:
                return {"error": f"Exa API error: {resp.status_code} - {resp.text[:500]}"}
            data = resp.json()
            results = []
            for r in data.get("results", []):
                results.append({
                    "title": r.get("title", ""),
                    "url": r.get("url", ""),
                    "published_date": r.get("publishedDate", ""),
                    "author": r.get("author", ""),
                    "text": r.get("text", "")[:2000],
                })
            return {"results": results, "total": len(results)}
        except httpx.TimeoutException:
            return {"error": "Exa API request timed out"}
        except Exception as e:
            return {"error": f"Exa API error: {str(e)}"}


async def _create_excel(data_str: str, title: str = "HIVE Report", filename: str = "report") -> dict:
    """Create a formatted Excel file from JSON data."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
    except ImportError:
        return {"error": "openpyxl not installed. Run: pip install openpyxl"}

    # Parse data
    try:
        data = json.loads(data_str) if isinstance(data_str, str) else data_str
    except json.JSONDecodeError:
        return {"error": "Invalid JSON data"}

    if not isinstance(data, list) or not data:
        return {"error": "Data must be a non-empty JSON array of objects"}

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Get headers
    headers = list(data[0].keys()) if isinstance(data[0], dict) else [f"Col {i}" for i in range(len(data[0]))]

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=str(header).upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "") if isinstance(row_data, dict) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.alignment = cell_alignment
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(data) + 2):
            try:
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Add title row
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Add timestamp
    ws.insert_rows(2)
    ts_cell = ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    ts_cell.font = Font(italic=True, color="808080")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))

    # Save
    desktop = os.path.join(os.path.expanduser("~"), "OneDrive", "Desktop")
    if not os.path.exists(desktop):
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    hive_output = os.path.join(os.path.expanduser("~"), ".hive", "output")
    os.makedirs(hive_output, exist_ok=True)

    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    fname = f"{filename}_{ts}.xlsx"
    filepath = os.path.join(hive_output, fname)
    desktop_path = os.path.join(desktop, fname)

    try:
        wb.save(filepath)
        wb.save(desktop_path)
        return {
            "status": "success",
            "file": filepath,
            "desktop_file": desktop_path,
            "rows": len(data),
            "columns": len(headers),
            "message": f"Excel saved to Desktop: {desktop_path}",
        }
    except Exception as e:
        return {"error": f"Failed to save: {str(e)}"}


# ─── System Tools ─────────────────────────────────────────────────────────────

async def _system_info() -> dict:
    """Get comprehensive system information."""
    import platform
    import sys
    
    info = {
        "os": platform.system(),
        "os_version": platform.version(),
        "architecture": platform.machine(),
        "python_version": sys.version,
        "python_executable": sys.executable,
        "hostname": platform.node(),
        "current_dir": os.getcwd(),
        "user": os.getenv("USERNAME") or os.getenv("USER", "unknown"),
    }
    
    # Try to get CPU/RAM info
    try:
        import psutil
        info["cpu_count"] = psutil.cpu_count()
        info["cpu_percent"] = psutil.cpu_percent(interval=0.1)
        ram = psutil.virtual_memory()
        info["ram_total_gb"] = round(ram.total / (1024**3), 2)
        info["ram_used_gb"] = round(ram.used / (1024**3), 2)
        info["ram_percent"] = ram.percent
        disk = psutil.disk_usage("/")
        info["disk_total_gb"] = round(disk.total / (1024**3), 2)
        info["disk_used_gb"] = round(disk.used / (1024**3), 2)
        info["disk_percent"] = round(disk.percent, 1)
    except ImportError:
        info["note"] = "Install psutil for detailed system info: pip install psutil"
    
    return info


async def _disk_usage(path: str) -> dict:
    """Check disk usage for a path."""
    p = Path(path).resolve()
    if not p.exists():
        return {"error": f"Path not found: {path}"}
    
    try:
        import psutil
        usage = psutil.disk_usage(str(p))
        return {
            "path": str(p),
            "total_gb": round(usage.total / (1024**3), 2),
            "used_gb": round(usage.used / (1024**3), 2),
            "free_gb": round(usage.free / (1024**3), 2),
            "percent": round(usage.percent, 1),
        }
    except ImportError:
        # Fallback for Windows
        if os.name == "nt":
            import ctypes
            free_bytes = ctypes.c_ulonglong(0)
            total_bytes = ctypes.c_ulonglong(0)
            ctypes.windll.kernel32.GetDiskFreeSpaceExW(
                str(p), None, ctypes.pointer(total_bytes), ctypes.pointer(free_bytes)
            )
            total = total_bytes.value
            free = free_bytes.value
            used = total - free
            return {
                "path": str(p),
                "total_gb": round(total / (1024**3), 2),
                "used_gb": round(used / (1024**3), 2),
                "free_gb": round(free / (1024**3), 2),
                "percent": round((used / total) * 100, 1) if total > 0 else 0,
            }
        return {"error": "Install psutil for disk info: pip install psutil"}


async def _ping(host: str) -> dict:
    """Ping a host to check connectivity."""
    param = "-n" if os.name == "nt" else "-c"
    timeout_param = "-w" if os.name == "nt" else "-W"
    
    proc = await asyncio.create_subprocess_exec(
        "ping", param, "4", timeout_param, "5", host,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await proc.communicate()
    
    output = stdout.decode(errors="replace")
    success = proc.returncode == 0
    
    # Extract time if possible
    import re
    time_match = re.search(r"time[<=](\d+\.?\d*)\s*ms", output)
    avg_time = float(time_match.group(1)) if time_match else None
    
    return {
        "host": host,
        "success": success,
        "average_ms": avg_time,
        "output": output[:500],
    }


async def _dns_lookup(domain: str) -> dict:
    """Look up DNS records for a domain."""
    import socket
    
    results = {"domain": domain, "records": {}}
    
    try:
        # A records (IPv4)
        ips = socket.getaddrinfo(domain, None, socket.AF_INET)
        results["records"]["A"] = list(set([addr[4][0] for addr in ips]))
    except Exception:
        pass
    
    try:
        # AAAA records (IPv6)
        ips = socket.getaddrinfo(domain, None, socket.AF_INET6)
        results["records"]["AAAA"] = list(set([addr[4][0] for addr in ips]))
    except Exception:
        pass
    
    try:
        # Get canonical name
        cname = socket.getfqdn(domain)
        if cname != domain:
            results["records"]["CNAME"] = cname
    except Exception:
        pass
    
    if not results["records"]:
        return {"error": f"Could not resolve {domain}"}
    
    return results


# ─── Clipboard Tools ──────────────────────────────────────────────────────────

async def _clipboard_get() -> dict:
    """Get clipboard contents."""
    try:
        import pyperclip
        text = pyperclip.paste()
        return {"content": text, "length": len(text)}
    except ImportError:
        # Fallback to platform-specific
        if os.name == "nt":
            proc = await asyncio.create_subprocess_exec(
                "powershell", "-command", "Get-Clipboard",
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, _ = await proc.communicate()
            text = stdout.decode(errors="replace").strip()
            return {"content": text, "length": len(text)}
        return {"error": "Install pyperclip: pip install pyperclip"}


async def _clipboard_set(text: str) -> dict:
    """Set clipboard contents."""
    try:
        import pyperclip
        pyperclip.copy(text)
        return {"success": True, "length": len(text)}
    except ImportError:
        if os.name == "nt":
            # Use base64 encoding to prevent injection
            import base64
            encoded = base64.b64encode(text.encode("utf-16-le")).decode("ascii")
            ps_script = f"$bytes = [Convert]::FromBase64String('{encoded}'); [System.Text.Encoding]::Unicode.GetString($bytes) | Set-Clipboard"
            proc = await asyncio.create_subprocess_exec(
                "powershell", "-command", ps_script,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            await proc.communicate()
            return {"success": True, "length": len(text)}
        return {"error": "Install pyperclip: pip install pyperclip"}


# ─── Environment Tools ────────────────────────────────────────────────────────

async def _get_env(name: str) -> dict:
    """Get an environment variable."""
    value = os.getenv(name)
    if value is None:
        return {"error": f"Environment variable '{name}' not found"}
    # Mask sensitive values
    sensitive = ["key", "secret", "token", "password", "pass"]
    if any(s in name.lower() for s in sensitive):
        masked = value[:4] + "****" + value[-4:] if len(value) > 8 else "****"
        return {"name": name, "value": masked, "masked": True}
    return {"name": name, "value": value, "masked": False}


async def _list_env() -> dict:
    """List all environment variables."""
    env_vars = []
    for key, value in sorted(os.environ.items()):
        sensitive = ["key", "secret", "token", "password", "pass"]
        masked = any(s in key.lower() for s in sensitive)
        env_vars.append({
            "name": key,
            "value": "****" if masked else value[:100],
            "masked": masked,
        })
    return {"variables": env_vars, "count": len(env_vars)}


# ─── Process Tools ────────────────────────────────────────────────────────────

async def _list_processes(filter_name: str = "") -> dict:
    """List running processes."""
    try:
        import psutil
        processes = []
        for proc in psutil.process_iter(["pid", "name", "cpu_percent", "memory_percent"]):
            try:
                info = proc.info
                if filter_name and filter_name.lower() not in info["name"].lower():
                    continue
                processes.append({
                    "pid": info["pid"],
                    "name": info["name"],
                    "cpu_percent": round(info["cpu_percent"] or 0, 1),
                    "memory_percent": round(info["memory_percent"] or 0, 1),
                })
            except (psutil.NoSuchProcess, psutil.AccessDenied):
                pass
        
        # Sort by memory usage
        processes.sort(key=lambda x: x["memory_percent"], reverse=True)
        return {"processes": processes[:50], "total": len(processes)}
    except ImportError:
        # Fallback to tasklist/pkill
        if os.name == "nt":
            proc = await asyncio.create_subprocess_exec(
                "tasklist", "/FO", "CSV",
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, _ = await proc.communicate()
            lines = stdout.decode(errors="replace").split("\n")[1:]
            processes = []
            for line in lines[:30]:
                parts = line.strip().split(",")
                if len(parts) >= 2:
                    name = parts[0].strip('"')
                    pid = parts[1].strip('"')
                    if filter_name and filter_name.lower() not in name.lower():
                        continue
                    processes.append({"pid": pid, "name": name})
            return {"processes": processes, "total": len(processes)}
        return {"error": "Install psutil: pip install psutil"}


async def _kill_process(target: str) -> dict:
    """Kill a process by PID or name."""
    try:
        import psutil
        
        # Try PID first
        try:
            pid = int(target)
            proc = psutil.Process(pid)
            name = proc.name()
            proc.terminate()
            return {"success": True, "killed": name, "pid": pid}
        except ValueError:
            # Not a PID, search by name
            killed = []
            for proc in psutil.process_iter(["pid", "name"]):
                try:
                    if target.lower() in proc.info["name"].lower():
                        proc.terminate()
                        killed.append({"pid": proc.info["pid"], "name": proc.info["name"]})
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
            if killed:
                return {"success": True, "killed": killed}
            return {"error": f"No process found matching '{target}'"}
    except ImportError:
        if os.name == "nt":
            proc = await asyncio.create_subprocess_exec(
                "taskkill", "/F", "/IM", f"{target}.exe" if not target.endswith(".exe") else target,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await proc.communicate()
            return {"success": proc.returncode == 0, "output": stdout.decode(errors="replace")}
        else:
            proc = await asyncio.create_subprocess_exec(
                "pkill", "-f", target,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
            )
            stdout, stderr = await proc.communicate()
            return {"success": proc.returncode == 0, "output": stdout.decode(errors="replace")}


# ─── Package Tools ────────────────────────────────────────────────────────────

async def _install_package(package: str) -> dict:
    """Install a Python package via pip."""
    proc = await asyncio.create_subprocess_exec(
        sys.executable, "-m", "pip", "install", package,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await proc.communicate()
    
    return {
        "success": proc.returncode == 0,
        "package": package,
        "output": (stdout.decode(errors="replace") + stderr.decode(errors="replace"))[:2000],
    }


async def _list_packages() -> dict:
    """List installed Python packages."""
    proc = await asyncio.create_subprocess_exec(
        sys.executable, "-m", "pip", "list", "--format=json",
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, _ = await proc.communicate()
    
    try:
        packages = json.loads(stdout.decode(errors="replace"))
        return {"packages": packages, "count": len(packages)}
    except json.JSONDecodeError:
        return {"error": "Failed to parse package list"}


# ─── Database Tools ───────────────────────────────────────────────────────────

async def _sqlite_query(db_path: str, query: str) -> dict:
    """Run a SQL query on a SQLite database."""
    import sqlite3
    
    p = Path(db_path).resolve()
    if not p.exists():
        return {"error": f"Database not found: {db_path}"}
    
    conn = None
    try:
        conn = sqlite3.connect(str(p))
        cursor = conn.execute(query)
        
        if query.strip().upper().startswith(("SELECT", "PRAGMA")):
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            rows = cursor.fetchall()
            return {
                "columns": columns,
                "rows": [dict(zip(columns, row)) for row in rows[:100]],
                "row_count": len(rows),
            }
        else:
            conn.commit()
            return {"success": True, "rows_affected": cursor.rowcount}
    except Exception as e:
        return {"error": str(e)}
    finally:
        if conn:
            conn.close()


async def _sqlite_tables(db_path: str) -> dict:
    """List tables in a SQLite database."""
    return await _sqlite_query(db_path, "SELECT name FROM sqlite_master WHERE type='table'")


# ─── HTTP Client Tools ────────────────────────────────────────────────────────

async def _http_request(method: str, url: str, headers: str = "", body: str = "") -> dict:
    """Make HTTP requests."""
    method = method.upper()
    
    # Parse headers
    parsed_headers = {}
    if headers:
        try:
            parsed_headers = json.loads(headers)
        except json.JSONDecodeError:
            for h in headers.split(","):
                if ":" in h:
                    k, v = h.split(":", 1)
                    parsed_headers[k.strip()] = v.strip()
    
    # Parse body
    parsed_body = None
    if body:
        try:
            parsed_body = json.loads(body)
        except json.JSONDecodeError:
            parsed_body = body
    
    async with httpx.AsyncClient(follow_redirects=True, timeout=15.0) as client:
        resp = await client.request(
            method, url,
            headers=parsed_headers,
            json=parsed_body if isinstance(parsed_body, dict) else None,
            content=parsed_body if isinstance(parsed_body, str) else None,
        )
        
        try:
            content = resp.json()
        except Exception:
            content = resp.text[:10000]
        
        return {
            "status_code": resp.status_code,
            "headers": dict(resp.headers),
            "content": content,
        }


# ─── Scheduling Tools ─────────────────────────────────────────────────────────

def _load_jobs() -> list:
    """Load jobs from file."""
    if os.path.exists(JOBS_FILE):
        with open(JOBS_FILE, "r") as f:
            return json.load(f)
    return []


def _save_jobs(jobs: list):
    """Save jobs to file."""
    os.makedirs(os.path.dirname(JOBS_FILE), exist_ok=True)
    with open(JOBS_FILE, "w") as f:
        json.dump(jobs, f, indent=2)


async def _set_reminder(message: str, time_str: str) -> dict:
    """Set a reminder or scheduled task."""
    import uuid
    
    # Parse time string
    now = datetime.now()
    time_lower = time_str.lower().strip()
    
    try:
        if time_lower.endswith("m") and len(time_lower) > 1:
            minutes = int(time_lower[:-1])
            trigger_at = now + timedelta(minutes=minutes)
        elif time_lower.endswith("h") and len(time_lower) > 1:
            hours = int(time_lower[:-1])
            trigger_at = now + timedelta(hours=hours)
        elif time_lower.endswith("d") and len(time_lower) > 1:
            days = int(time_lower[:-1])
            trigger_at = now + timedelta(days=days)
        elif "-" in time_str and ":" in time_str:
            trigger_at = datetime.strptime(time_str, "%Y-%m-%d %H:%M")
        else:
            return {"error": "Invalid time format. Use: 5m, 2h, 3d, or 2024-01-01 10:00"}
    except (ValueError, IndexError):
        return {"error": "Invalid time format. Use: 5m, 2h, 3d, or 2024-01-01 10:00"}
    
    job_id = str(uuid.uuid4())[:8]
    job = {
        "id": job_id,
        "message": message,
        "trigger_at": trigger_at.isoformat(),
        "created_at": now.isoformat(),
        "status": "pending",
    }
    
    jobs = _load_jobs()
    jobs.append(job)
    _save_jobs(jobs)
    
    return {
        "success": True,
        "job_id": job_id,
        "message": message,
        "trigger_at": trigger_at.strftime("%Y-%m-%d %H:%M:%S"),
        "in_minutes": round((trigger_at - now).total_seconds() / 60, 1),
    }


async def _list_jobs() -> dict:
    """List scheduled jobs/reminders."""
    jobs = _load_jobs()
    now = datetime.now()
    
    active = []
    for job in jobs:
        trigger_at = datetime.fromisoformat(job["trigger_at"])
        if trigger_at > now and job["status"] == "pending":
            job["time_until"] = str(trigger_at - now).split(".")[0]
            active.append(job)
    
    return {"jobs": active, "count": len(active)}


async def _cancel_job(job_id: str) -> dict:
    """Cancel a scheduled job."""
    jobs = _load_jobs()
    for job in jobs:
        if job["id"] == job_id:
            job["status"] = "cancelled"
            _save_jobs(jobs)
            return {"success": True, "cancelled": job["message"]}
    return {"error": f"Job not found: {job_id}"}


# ─── Encryption Tools ─────────────────────────────────────────────────────────

async def _hash_text(text: str, algorithm: str = "sha256") -> dict:
    """Hash text using various algorithms."""
    algorithms = {
        "md5": hashlib.md5,
        "sha1": hashlib.sha1,
        "sha256": hashlib.sha256,
        "sha512": hashlib.sha512,
    }
    
    algo_func = algorithms.get(algorithm.lower())
    if not algo_func:
        return {"error": f"Unknown algorithm: {algorithm}. Use: md5, sha1, sha256, sha512"}
    
    hash_obj = algo_func(text.encode("utf-8"))
    return {
        "algorithm": algorithm.upper(),
        "hash": hash_obj.hexdigest(),
        "length": len(hash_obj.hexdigest()),
    }


async def _hash_file(path: str, algorithm: str = "sha256") -> dict:
    """Hash a file using various algorithms."""
    p = Path(path).resolve()
    if not p.exists():
        return {"error": f"File not found: {path}"}
    
    algorithms = {
        "md5": hashlib.md5,
        "sha1": hashlib.sha1,
        "sha256": hashlib.sha256,
        "sha512": hashlib.sha512,
    }
    
    algo_func = algorithms.get(algorithm.lower())
    if not algo_func:
        return {"error": f"Unknown algorithm: {algorithm}. Use: md5, sha1, sha256, sha512"}
    
    hash_obj = algo_func()
    with open(p, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            hash_obj.update(chunk)
    
    return {
        "file": str(p),
        "algorithm": algorithm.upper(),
        "hash": hash_obj.hexdigest(),
        "size_bytes": p.stat().st_size,
    }


async def _base64_encode(text: str) -> dict:
    """Encode text to base64."""
    encoded = base64.b64encode(text.encode("utf-8")).decode("utf-8")
    return {"encoded": encoded, "length": len(encoded)}


async def _base64_decode(text: str) -> dict:
    """Decode base64 to text."""
    try:
        decoded = base64.b64decode(text.encode("utf-8")).decode("utf-8")
        return {"decoded": decoded, "length": len(decoded)}
    except Exception as e:
        return {"error": f"Invalid base64: {e}"}


# ─── Archive Tools ────────────────────────────────────────────────────────────

async def _zip_files(source: str, output: str) -> dict:
    """Create a zip archive from files/directories."""
    source_path = Path(source).resolve()
    output_path = Path(output).resolve()
    
    if not source_path.exists():
        return {"error": f"Source not found: {source}"}
    
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    try:
        with zipfile.ZipFile(str(output_path), "w", zipfile.ZIP_DEFLATED) as zf:
            if source_path.is_file():
                zf.write(source_path, source_path.name)
            elif source_path.is_dir():
                for file in source_path.rglob("*"):
                    if file.is_file():
                        arcname = file.relative_to(source_path.parent)
                        zf.write(file, arcname)
        
        return {
            "success": True,
            "archive": str(output_path),
            "size_bytes": output_path.stat().st_size,
            "source": str(source_path),
        }
    except Exception as e:
        return {"error": str(e)}


async def _unzip_file(source: str, output: str = "") -> dict:
    """Extract a zip archive safely (prevents Zip Slip)."""
    source_path = Path(source).resolve()
    
    if not source_path.exists():
        return {"error": f"Zip file not found: {source}"}
    
    if not output:
        output = str(source_path.parent / source_path.stem)
    
    output_path = Path(output).resolve()
    output_path.mkdir(parents=True, exist_ok=True)
    
    try:
        with zipfile.ZipFile(str(source_path), "r") as zf:
            # Security: prevent Zip Slip path traversal
            for info in zf.infolist():
                target = (output_path / info.filename).resolve()
                if not str(target).startswith(str(output_path)):
                    return {"error": f"Zip Slip detected: {info.filename} tries to escape output directory"}
            
            zf.extractall(str(output_path))
        
        files = list(output_path.rglob("*"))
        return {
            "success": True,
            "extracted_to": str(output_path),
            "files_count": len([f for f in files if f.is_file()]),
            "dirs_count": len([f for f in files if f.is_dir()]),
        }
    except Exception as e:
        return {"error": str(e)}


async def _list_zip(source: str) -> dict:
    """List contents of a zip archive."""
    source_path = Path(source).resolve()
    
    if not source_path.exists():
        return {"error": f"Zip file not found: {source}"}
    
    try:
        with zipfile.ZipFile(str(source_path), "r") as zf:
            files = []
            for info in zf.infolist():
                files.append({
                    "name": info.filename,
                    "size": info.file_size,
                    "compressed": info.compress_size,
                    "ratio": round((1 - info.compress_size / info.file_size) * 100, 1) if info.file_size > 0 else 0,
                })
            
            return {
                "archive": str(source_path),
                "files": files,
                "total_files": len(files),
                "total_size": sum(f["size"] for f in files),
                "compressed_size": sum(f["compressed"] for f in files),
            }
    except Exception as e:
        return {"error": str(e)}


# ─── Browser Tool Implementations ──────────────────────────────────

# Single source of truth for interactive-element discovery.
# Tags each discovered element with data-hive-index so browser_click /
# browser_type resolve the EXACT element browser_inspect reported —
# previously each tool rebuilt its own list and the indices didn't match
# (e.g. hidden CSRF inputs shifted click targets).
_INSPECT_TAG_JS = """() => {
    // Clear stale tags from a previous inspect
    document.querySelectorAll('[data-hive-index]').forEach(el => el.removeAttribute('data-hive-index'));

    const selectors = [
        'input', 'button', 'a', 'select', 'textarea',
        '[role="button"]', '[role="link"]', '[role="tab"]',
        '[role="checkbox"]', '[role="radio"]', '[role="combobox"]',
        '[role="switch"]', '[role="menuitem"]', '[role="option"]',
        '[onclick]', '[contenteditable]',
        '[data-testid]',
        'form[role="search"]',
    ];
    // NOTE: querySelectorAll never returns the same node twice, so no
    // dedup is needed. (The old text-based dedup key collapsed distinct
    // empty inputs — e.g. username + password fields — into one entry.)
    const els = document.querySelectorAll(selectors.join(', '));
    const results = [];
    for (const el of els) {
        const tag = el.tagName.toLowerCase();
        const type = el.type || '';

        // Skip invisible/hidden elements (hidden inputs, offscreen nodes)
        const rect = el.getBoundingClientRect();
        if (rect.width === 0 && rect.height === 0) continue;
        if (type === 'hidden') continue;

        const text = el.textContent?.trim().substring(0, 60) || '';
        const placeholder = el.placeholder || '';
        const ariaLabel = el.getAttribute('aria-label') || '';
        const role = el.getAttribute('role') || '';
        const testId = el.getAttribute('data-testid') || '';
        const required = el.required || false;
        const disabled = el.disabled || false;

        let selector = '';
        if (testId) selector = '[data-testid="' + testId + '"]';
        else if (el.id) selector = '#' + el.id;
        else if (el.name) selector = tag + '[name="' + el.name + '"]';

        el.setAttribute('data-hive-index', String(results.length));

        results.push({
            index: results.length,
            tag: tag,
            type: type,
            text: text,
            placeholder: placeholder,
            ariaLabel: ariaLabel,
            role: role,
            testId: testId,
            required: required,
            disabled: disabled,
            selector: selector,
        });
        if (results.length >= 50) break;
    }
    return results;
}"""


async def _get_indexed_element(page, index: int):
    """Resolve an element by the index browser_inspect reported.

    Uses the data-hive-index attribute stamped during inspect; re-tags the
    page if the attribute is gone (e.g. after a re-render or navigation).
    """
    el = await page.query_selector(f'[data-hive-index="{index}"]')
    if el is None:
        await page.evaluate(_INSPECT_TAG_JS)
        el = await page.query_selector(f'[data-hive-index="{index}"]')
    return el


async def _browser_open(url: str) -> dict:
    """Open a URL in the browser."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        # Wait extra time for React/SPA to render
        await page.wait_for_timeout(3000)
        title = await page.title()
        return {"status": "opened", "url": page.url, "title": title}
    except Exception as e:
        return {"error": str(e)}


async def _browser_click(selector: str = "", index: int = -1) -> dict:
    """Click an element by index (from browser_inspect) or CSS selector/text."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        if index >= 0:
            # Resolve the exact element browser_inspect reported at this index
            el = await _get_indexed_element(page, index)
            if el is None:
                return {"error": f"Index {index} not found on page (run browser_inspect again)"}
            try:
                await el.scroll_into_view_if_needed(timeout=3000)
            except Exception:
                pass
            # Try normal click first, fall back to JavaScript click
            try:
                await el.click(timeout=5000)
            except Exception:
                # JavaScript click fallback (bypasses bot detection)
                await el.evaluate("(el) => el.click()")
        elif selector:
            try:
                await page.click(selector, timeout=5000)
            except Exception:
                # JavaScript click fallback
                try:
                    el = await page.wait_for_selector(selector, timeout=5000)
                    await el.evaluate("(el) => el.click()")
                except Exception:
                    # Try as visible text
                    try:
                        await page.get_by_text(selector, exact=False).first.click(timeout=5000)
                    except Exception as final_err:
                        return {"error": f"Could not click '{selector}': {final_err}"}
        else:
            return {"error": "Provide either index or selector"}
        await page.wait_for_load_state("domcontentloaded", timeout=10000)
        return {"status": "clicked", "url": page.url}
    except Exception as e:
        return {"error": str(e)}


async def _type_like_human(page, el, text: str, press_enter: bool) -> None:
    """Type into an element using REAL key events (via CDP), like a person.

    This is far less detectable than setting el.value in JS. Falls back to the
    native-value-setter approach only if real typing didn't register (some
    exotic controlled inputs).
    """
    try:
        await el.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    # Focus and clear any existing value with a real select-all + delete
    try:
        await el.click(timeout=3000)
    except Exception:
        try:
            await el.focus()
        except Exception:
            pass
    try:
        await page.keyboard.press("Control+A")
        await page.keyboard.press("Delete")
    except Exception:
        pass

    typed_ok = True
    try:
        # press_sequentially issues real keydown/keypress/input/keyup events
        await el.press_sequentially(text, delay=45)
    except Exception:
        try:
            await el.type(text, delay=45)
        except Exception:
            typed_ok = False

    # Verify the value took; fall back to native setter for stubborn inputs
    if not typed_ok:
        try:
            await el.evaluate("""(el, text) => {
                el.focus();
                const proto = el.tagName === 'TEXTAREA'
                    ? window.HTMLTextAreaElement.prototype
                    : window.HTMLInputElement.prototype;
                const setter = Object.getOwnPropertyDescriptor(proto, 'value');
                if (setter && setter.set) setter.set.call(el, text);
                else el.value = text;
                el.dispatchEvent(new Event('input', { bubbles: true }));
                el.dispatchEvent(new Event('change', { bubbles: true }));
            }""", text)
        except Exception:
            pass

    if press_enter:
        try:
            await el.press("Enter")
        except Exception:
            await page.keyboard.press("Enter")


async def _browser_type(selector: str = "", text: str = "", press_enter: bool = False, index: int = -1, sensitive_data: dict = None) -> dict:
    """Type text into an input field by index (from browser_inspect) or CSS selector."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        
        # Replace ${key} placeholders with sensitive_data values
        actual_text = text
        if sensitive_data:
            for key, value in sensitive_data.items():
                actual_text = actual_text.replace(f"${{{key}}}", str(value))
        
        if index >= 0:
            # Resolve the exact element browser_inspect reported at this index
            el = await _get_indexed_element(page, index)
            if el is None:
                return {"error": f"Index {index} not found on page (run browser_inspect again)"}
            await _type_like_human(page, el, actual_text, press_enter)
        elif selector:
            el = await page.wait_for_selector(selector, timeout=5000)
            await _type_like_human(page, el, actual_text, press_enter)
        else:
            return {"error": "Provide either index or selector"}
        await page.wait_for_load_state("domcontentloaded", timeout=10000)
        return {"status": "typed", "text": "***" if sensitive_data else actual_text, "url": page.url}
    except Exception as e:
        return {"error": str(e)}


async def _browser_wait(selector: str, timeout: int = 10) -> dict:
    """Wait for an element to appear."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        await page.wait_for_selector(selector, timeout=timeout * 1000)
        return {"status": "found", "selector": selector}
    except Exception as e:
        return {"error": f"Timeout waiting for {selector}: {e}"}


async def _browser_screenshot(filename: str = "screenshot.png") -> dict:
    """Take a screenshot and save to HIVE screenshots directory."""
    from hive.browser.pool import get_pool
    from hive.config import BROWSER_SCREENSHOT_DIR, ensure_dirs
    ensure_dirs()
    pool = get_pool()
    try:
        page = await pool.get_page()
        filepath = BROWSER_SCREENSHOT_DIR / filename
        await page.screenshot(path=str(filepath), full_page=False)
        return {"status": "saved", "path": str(filepath)}
    except Exception as e:
        return {"error": str(e)}


async def _browser_read() -> dict:
    """Read the text content of the current page."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        title = await page.title()
        text = await page.inner_text("body")
        # Limit to 5000 chars to avoid token overflow
        if len(text) > 5000:
            text = text[:5000] + "\n... (truncated)"
        return {"title": title, "url": page.url, "text": text}
    except Exception as e:
        return {"error": str(e)}


async def _browser_inspect() -> dict:
    """List all interactive elements on the page with index numbers."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        # Discover interactive elements and stamp data-hive-index on each so
        # browser_click / browser_type resolve the exact same elements.
        elements = await page.evaluate(_INSPECT_TAG_JS)
        return {
            "elements": elements,
            "count": len(elements),
            "url": page.url,
            "title": await page.title(),
        }
    except Exception as e:
        return {"error": str(e)}


async def _browser_back() -> dict:
    """Navigate back in browser history."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        page = await pool.get_page()
        await page.go_back(wait_until="domcontentloaded", timeout=15000)
        title = await page.title()
        return {"status": "navigated_back", "url": page.url, "title": title}
    except Exception as e:
        return {"error": str(e)}


async def _browser_close() -> dict:
    """Close the browser session."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        await pool.close_session()
        return {"status": "closed"}
    except Exception as e:
        return {"error": str(e)}


async def _browser_session_save(name: str) -> dict:
    """Save current browser session to disk."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        path = await pool.save_session(name)
        return {"status": "saved", "name": name, "path": path}
    except Exception as e:
        return {"error": str(e)}


async def _browser_session_load(name: str) -> dict:
    """Load a saved browser session."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        success = await pool.load_session(name)
        if success:
            page = await pool.get_page()
            title = await page.title()
            return {"status": "loaded", "name": name, "url": page.url, "title": title}
        else:
            return {"error": f"Session '{name}' not found"}
    except Exception as e:
        return {"error": str(e)}


async def _browser_list_sessions() -> dict:
    """List all saved browser sessions."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        sessions = await pool.list_sessions()
        return {"sessions": sessions, "count": len(sessions)}
    except Exception as e:
        return {"error": str(e)}


async def _browser_delete_session(name: str) -> dict:
    """Delete a saved browser session."""
    from hive.browser.pool import get_pool
    pool = get_pool()
    try:
        success = await pool.delete_session(name)
        if success:
            return {"status": "deleted", "name": name}
        else:
            return {"error": f"Session '{name}' not found"}
    except Exception as e:
        return {"error": str(e)}


# Email verification state
_inbox_data = {"id": None, "address": None, "api_key": None}


async def _browser_create_inbox() -> dict:
    """Create a disposable email inbox for verification."""
    import httpx
    import os
    
    api_key = os.getenv("MINUTEMAIL_API_KEY") or os.getenv("MAILSINK_API_KEY")
    if not api_key:
        return {"error": "MINUTEMAIL_API_KEY or MAILSINK_API_KEY not set in .env"}
    
    try:
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.minutemail.co/mailboxes",
                headers={"Authorization": f"Bearer {api_key}"},
                json={"ttl": 10},
                timeout=15
            )
            if response.status_code == 200:
                data = response.json()
                _inbox_data["id"] = data.get("id")
                _inbox_data["address"] = data.get("address")
                _inbox_data["api_key"] = api_key
                return {
                    "status": "created",
                    "email": data.get("address"),
                    "inbox_id": data.get("id"),
                    "expires_in": "10 minutes"
                }
            else:
                return {"error": f"API error: {response.status_code} - {response.text}"}
    except Exception as e:
        return {"error": str(e)}


async def _browser_wait_for_code(timeout: int = 30, sender_hint: str = "") -> dict:
    """Wait for a verification code — disposable inbox first, then the
    user's personal email over IMAP (if configured)."""
    import httpx
    import re
    import time

    if not _inbox_data["id"] or not _inbox_data["api_key"]:
        # No disposable inbox active — try the user's own inbox via IMAP
        from hive.browser.email_codes import resolve_imap_config, wait_for_code_imap
        if resolve_imap_config():
            return await wait_for_code_imap(timeout=max(timeout, 60), sender_hint=sender_hint)
        return {
            "error": (
                "No inbox available. Either call browser_create_inbox first (disposable), "
                "or configure your personal inbox: store an app password via "
                "vault_store_credential(site='imap', ...) or set HIVE_IMAP_USER / "
                "HIVE_IMAP_PASSWORD in .env."
            )
        }
    
    try:
        start_time = time.time()
        async with httpx.AsyncClient() as client:
            while time.time() - start_time < timeout:
                response = await client.get(
                    f"https://api.minutemail.co/mailboxes/{_inbox_data['id']}/mails",
                    headers={"Authorization": f"Bearer {_inbox_data['api_key']}"},
                    timeout=10
                )
                if response.status_code == 200:
                    data = response.json()
                    items = data.get("items", [])
                    if items:
                        body = items[0].get("body", "") or items[0].get("text", "")
                        # Extract 4-8 digit code
                        match = re.search(r'\b(\d{4,8})\b', body)
                        if match:
                            return {"status": "received", "code": match.group(1)}
                await asyncio.sleep(2)
        return {"error": f"No code received within {timeout} seconds"}
    except Exception as e:
        return {"error": str(e)}


async def _browser_use_task(task: str, headless: bool = False, max_steps: int = 30) -> dict:
    """Automate browser tasks using Browser Use with cloned Chrome profile."""
    try:
        from hive.agents.workers.browser_use_worker import _run_browser_use_task
        result = await _run_browser_use_task(task, headless=headless, max_steps=max_steps)
        return {"status": "completed", "output": result}
    except ImportError:
        return {"error": "browser-use library not installed. Run: pip install browser-use"}
    except Exception as e:
        return {"error": f"Browser task failed: {str(e)}"}


# ─── Vault Tool Implementations ───────────────────────────────────

async def _vault_store_credential(site: str, username: str, password: str) -> dict:
    from hive.browser.vault import store_credential
    cred_id = store_credential(site, username, password)
    return {"status": "stored", "id": cred_id, "site": site, "username": username}


async def _vault_list_credentials() -> dict:
    from hive.browser.vault import list_credentials
    creds = list_credentials()
    return {"credentials": creds, "count": len(creds)}


async def _vault_store_card(label: str, number: str, expiry: str, cvv: str,
                             name: str = "", billing_zip: str = "") -> dict:
    from hive.browser.vault import store_card
    card_id = store_card(label, number, expiry, cvv, name, billing_zip)
    return {"status": "stored", "id": card_id, "label": label, "last4": number[-4:]}


async def _vault_list_cards() -> dict:
    from hive.browser.vault import list_cards
    cards = list_cards()
    return {"cards": cards, "count": len(cards)}


async def _browser_signup(task: str, url: str = None, email: str = None, password: str = None) -> dict:
    from hive.browser.signup import run_signup
    return await run_signup(task, url=url, email=email, password=password)


async def _browser_checkout(task: str, url: str = None, amount: float = None,
                             card_id: str = None, confirm: bool = False) -> dict:
    from hive.browser.checkout import run_checkout, confirm_checkout
    if confirm:
        return await confirm_checkout(task, url=url, amount=amount, card_id=card_id)
    return await run_checkout(task, url=url, amount=amount, card_id=card_id)


async def _browser_google_login(email: str = None) -> dict:
    from hive.browser.google_login import run_google_login
    return await run_google_login(email=email)


async def _browser_oauth(platform: str = "github") -> dict:
    from hive.browser.oauth import start_oauth
    return await start_oauth(platform)
